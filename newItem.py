# -- listPrice.py file

# Map each category_code to its divisor
markup_dict = {
    "25D": 0.250, "25B": 0.300, "27Q": 0.700, "21F": 0.300, "26G": 0.300, "27H": 0.120, "33B": 0.201,
    "44A": 0.240, "45A": 0.380, "45F": 0.200, "46R": 0.600, "50I": 0.100, "50O": 0.700, "50P": 0.200
}


def markup(category_code, tpp_value):
    try:
        divisor = markup_dict[category_code]
        tpp_value = float(tpp_value)
        list_price = round((1.15 * tpp_value) / divisor, 4)
        return list_price
    except KeyError:
        print(f"Undefined Item Category: '{category_code}'")
        return None
    except (ValueError, TypeError):
        print("Invalid TPP value.")
        return None


# --  --  --  --  --  --  --  --  --  --  --  --  
# ----- Begin Part Number Creation Program -----
# --  --  --  --  --  --  --  --  --  --  --  --  

from listPrice import markup # import function from above
import openpyxl as op
from datetime import date
import sys

# Excel File Path
newPart_form = "C:/Users/NewPartNumber/NewPartNumber_Form.xlsx"
newPart_table = "C:/Users/NewPartNumber/NewPartNumber_Table.xlsx"


#Load Form Workbook (openpyxl)
wb_form = op.load_workbook(newPart_form)
form = wb_form["newPart"]


#Form Variables
vnd_name  = form["B2"].value.strip()
vnd_id    = form["B3"].value.upper().strip()
sku       = str(form["B4"].value).strip()
detail    = form["B5"].value.strip()
tpp       = float(form["B10"].value)
cat_code  = form["B11"].value
site      = form["B12"].value.upper()
request   = form["B13"].value
prod_grp  = form["B14"].value


#Date Variable
today = date.today()


#Item Description
item_descr = (f"{vnd_name},#{sku},{detail}").upper()


#List Price Function
list_price = markup(cat_code, tpp)

if list_price is None:
    sys.exit()
else:
    print(f"""Form accessed, markup applied: {today}""")


#Load Table Workbook (openpyxl)
wb_table = op.load_workbook(newPart_table)
table = wb_table["699_Table"]


#Select Next Blank Row in Table
next_row = 1
while table.cell(row=next_row, column=1).value is not None:
    next_row += 1


#Select Previous Row (for previous part number)
prev_pn_cell = table.cell(row= next_row - 1, column= 1).value
previous_pn = int(prev_pn_cell)


#Create New Part Number
new_pn = previous_pn + 1


#Fill Next Blank Row on Table
table.cell(row = next_row, column = 1,  value = new_pn)
table.cell(row = next_row, column = 2,  value = sku)
table.cell(row = next_row, column = 3,  value = item_descr)
table.cell(row = next_row, column = 4,  value = today)
table.cell(row = next_row, column = 5,  value = site)
table.cell(row = next_row, column = 6,  value = tpp)
table.cell(row = next_row, column = 7,  value = cat_code)
table.cell(row = next_row, column = 11, value = list_price)
table.cell(row = next_row, column = 12, value = vnd_id)

wb_table.save(newPart_table)
wb_table.close()


#Print in Terminal for Data Entry
print(f"""
New PN:   {new_pn}
VNDID:    {vnd_id}
Item:     {item_descr}
TPP:      {tpp}
List:     {list_price}
Request:  {request}
Site:     {site}

Prod Grp: {prod_grp}
""")
